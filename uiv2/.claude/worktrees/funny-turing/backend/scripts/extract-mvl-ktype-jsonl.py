#!/usr/bin/env python3
"""
Extract a compact, backend-friendly MVL dataset from the decrypted eBay MVL workbook.

Input:
  exports/DE_MVL_2025_10.decrypted.xlsx  (sheet: "DE_MVL_2025_10")

Output:
  exports/DE_MVL_2025_10.compact.jsonl

Each JSON line:
  {
    "k": 19842,
    "make": "Audi",
    "model": "A4",
    "type": "2.0 TDI",
    "platform": "8EC",
    "period": "2006/06-2008/06",
    "engine": "1968 ccm, 125 KW, 170 PS",
    "hsn_tsn": "0588|AFK"
  }
"""

from __future__ import annotations

import json
from pathlib import Path

from openpyxl import load_workbook


ROOT = Path(__file__).resolve().parents[2]
SRC = ROOT / "exports" / "DE_MVL_2025_10.decrypted.xlsx"
OUT = ROOT / "exports" / "DE_MVL_2025_10.compact.jsonl"


def s(v) -> str:
    return "" if v is None else str(v).strip()


def main() -> None:
    if not SRC.exists():
        raise SystemExit(f"Missing decrypted workbook: {SRC} (run decrypt script first)")

    wb = load_workbook(filename=str(SRC), data_only=True, read_only=True)
    if "DE_MVL_2025_10" not in wb.sheetnames:
        raise SystemExit(f"Sheet DE_MVL_2025_10 not found. Sheets: {wb.sheetnames}")
    ws = wb["DE_MVL_2025_10"]

    # Use iter_rows(values_only=True) for performance (cell-by-cell is slow for 55k rows).
    rows_iter = ws.iter_rows(min_row=1, max_row=1, values_only=True)
    headers = [s(h) for h in next(rows_iter)]
    idx0 = {h: i for i, h in enumerate(headers) if h}  # 0-based

    required = [
        "K-Type",
        "Marke_Make_EN",
        "Modell_Model_EN",
        "Typ_Type_EN",
        "Plattform_Platform_EN",
        "Baujahr_ProductionPeriod_EN",
        "Motor_Engine_EN",
        "HSN_TSN_nur_zur_Hilfe",
    ]
    missing = [h for h in required if h not in idx0]
    if missing:
        raise SystemExit(f"Missing expected columns: {missing}. Got: {headers}")

    out_f = OUT.open("w", encoding="utf-8")
    read_rows = 0
    wrote = 0

    def col(name: str) -> int:
        return idx0[name]

    data_iter = ws.iter_rows(min_row=2, values_only=True)
    for row in data_iter:
        read_rows += 1
        if read_rows % 5000 == 0:
            print("progress_rows ->", read_rows, "wrote ->", wrote)
        k_raw = s(row[col("K-Type")])
        if not k_raw:
            continue
        try:
            k = int(k_raw)
        except ValueError:
            continue

        rec = {
            "k": k,
            "make": s(row[col("Marke_Make_EN")]),
            "model": s(row[col("Modell_Model_EN")]),
            "type": s(row[col("Typ_Type_EN")]),
            "platform": s(row[col("Plattform_Platform_EN")]),
            "period": s(row[col("Baujahr_ProductionPeriod_EN")]),
            "engine": s(row[col("Motor_Engine_EN")]),
            "hsn_tsn": s(row[col("HSN_TSN_nur_zur_Hilfe")]),
        }
        out_f.write(json.dumps(rec, ensure_ascii=False) + "\n")
        wrote += 1

    out_f.close()

    print("read_rows ->", read_rows)
    print("wrote_rows ->", wrote)
    print("out ->", OUT)


if __name__ == "__main__":
    main()

