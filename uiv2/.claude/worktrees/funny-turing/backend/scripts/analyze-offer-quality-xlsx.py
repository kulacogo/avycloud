#!/usr/bin/env python3
"""
Analyze an eBay "Bericht zur Angebotsqualität" XLSX (multi-sheet) and export:
- Per-sheet overview
- Per-category offer table metrics (missing fields, counts)
- Full offer rows (long format CSV)
- Google Shopping rejected products table

This is a read-only, deterministic analysis tool (no network).

Usage:
  python3 backend/scripts/analyze-offer-quality-xlsx.py \\
    --in "/path/to/Bericht....xlsx" \\
    --out-dir "/path/to/exports/offer-quality/20260107-0332"
"""

from __future__ import annotations

import argparse
import csv
import json
import os
import re
from collections import Counter, defaultdict
from dataclasses import dataclass
from datetime import datetime
from pathlib import Path
from typing import Any, Dict, List, Optional, Tuple

from openpyxl import load_workbook


def safe_str(v: Any) -> str:
    if v is None:
        return ""
    return str(v).strip()


def norm_space(s: str) -> str:
    return re.sub(r"\s+", " ", safe_str(s)).strip()


def lower(s: Any) -> str:
    return safe_str(s).strip().lower()


def is_missing(v: Any) -> bool:
    s = lower(v)
    if not s:
        return True
    if s == "keine angabe":
        return True
    return False


def is_present_checkmark(v: Any) -> bool:
    # openpyxl returns the check mark as a string
    return safe_str(v) in ("✔", "✓", "✅")


def parse_int(v: Any) -> Optional[int]:
    if v is None:
        return None
    if isinstance(v, (int,)):
        return int(v)
    if isinstance(v, float):
        if v.is_integer():
            return int(v)
        # sometimes counts are exported as floats; round safely
        return int(round(v))
    s = safe_str(v)
    if not s:
        return None
    m = re.search(r"-?\d+", s.replace(".", "").replace(",", ""))
    if not m:
        return None
    try:
        return int(m.group(0))
    except Exception:
        return None


def write_json(path: Path, obj: Any) -> None:
    path.write_text(json.dumps(obj, ensure_ascii=False, indent=2), encoding="utf-8")


def write_csv(path: Path, rows: List[Dict[str, Any]], headers: List[str]) -> None:
    with path.open("w", encoding="utf-8", newline="") as f:
        writer = csv.DictWriter(f, fieldnames=headers, extrasaction="ignore")
        writer.writeheader()
        for row in rows:
            writer.writerow({h: row.get(h, "") for h in headers})


@dataclass
class Table:
    header_row: int
    last_row: int
    last_col: int
    headers: List[str]
    rows: List[List[Any]]


def find_row_with_cell_value(ws, value: str, *, col: Optional[int] = None, max_rows: Optional[int] = None) -> Optional[int]:
    target = value.strip()
    max_r = ws.max_row if max_rows is None else min(ws.max_row, max_rows)
    if col is not None:
        for r in range(1, max_r + 1):
            v = ws.cell(row=r, column=col).value
            if isinstance(v, str) and v.strip() == target:
                return r
        return None
    for r in range(1, max_r + 1):
        for c in range(1, ws.max_column + 1):
            v = ws.cell(row=r, column=c).value
            if isinstance(v, str) and v.strip() == target:
                return r
    return None


def infer_last_col_from_header(ws, header_row: int) -> int:
    last_col = 0
    for c in range(1, ws.max_column + 1):
        v = norm_space(ws.cell(row=header_row, column=c).value)
        if v:
            last_col = c
    return last_col or ws.max_column


def infer_last_row_from_col(ws, start_row: int, title_col: int, *, max_empty_run: int = 8) -> int:
    empty_run = 0
    last = start_row
    for r in range(start_row, ws.max_row + 1):
        v = ws.cell(row=r, column=title_col).value
        if safe_str(v):
            empty_run = 0
            last = r
            continue
        empty_run += 1
        if empty_run >= max_empty_run:
            break
    return last


def extract_table(ws, header_row: int, *, title_col: int = 2) -> Table:
    last_col = infer_last_col_from_header(ws, header_row)
    last_row = infer_last_row_from_col(ws, header_row + 1, title_col)

    headers = [norm_space(ws.cell(row=header_row, column=c).value) for c in range(1, last_col + 1)]
    # Replace empty headers with deterministic placeholder
    headers = [h if h else f"COL_{i+1}" for i, h in enumerate(headers)]

    rows = []
    for r in range(header_row + 1, last_row + 1):
        first = ws.cell(row=r, column=title_col).value
        if not safe_str(first):
            continue
        rows.append([ws.cell(row=r, column=c).value for c in range(1, last_col + 1)])

    return Table(
        header_row=header_row,
        last_row=last_row,
        last_col=last_col,
        headers=headers,
        rows=rows,
    )


def sheet_type(name: str) -> str:
    if name.strip() == "Zusammenfassung":
        return "summary"
    if name.strip() == "Ratgeber":
        return "guide"
    if name.strip().lower().startswith("google shopping: abgelehnte produkte"):
        return "google_shopping_rejected"
    return "category"


def analyze_summary_sheet(ws) -> Dict[str, Any]:
    # Extract blocks where a category line and "X Angebote können optimiert werden" exist on the same row.
    blocks = []
    rx = re.compile(r"(\d+)\s+Angebote\s+können\s+optimiert\s+werden", re.IGNORECASE)
    for r in range(1, ws.max_row + 1):
        cat = safe_str(ws.cell(row=r, column=2).value)
        status = safe_str(ws.cell(row=r, column=6).value)
        if not cat or not status:
            continue
        m = rx.search(status)
        if not m:
            continue
        blocks.append(
            {
                "row": r,
                "category": norm_space(cat),
                "optimizable_offers": int(m.group(1)),
                "status_raw": status,
            }
        )

    return {
        "sheet": ws.title,
        "type": "summary",
        "optimizable_blocks": blocks,
        "optimizable_total": sum(b["optimizable_offers"] for b in blocks),
        "unique_categories": len({b["category"] for b in blocks}),
    }


def analyze_google_shopping_sheet(ws) -> Dict[str, Any]:
    header_row = find_row_with_cell_value(ws, "Google-Problem", max_rows=200)
    if not header_row:
        return {"sheet": ws.title, "type": "google_shopping_rejected", "table": None}
    table = extract_table(ws, header_row, title_col=4)  # Angebotstitel is usually col 4
    # Map rows
    entries = []
    for row in table.rows:
        item = {}
        for i, h in enumerate(table.headers):
            item[h] = safe_str(row[i])
        entries.append(item)

    # normalize common keys
    def pick(item, *keys):
        for k in keys:
            if k in item and safe_str(item[k]):
                return safe_str(item[k])
        return ""

    issues = Counter()
    for item in entries:
        issues[pick(item, "Google-Problem", "Google Problem")] += 1

    return {
        "sheet": ws.title,
        "type": "google_shopping_rejected",
        "table": {
            "header_row": table.header_row,
            "rows": len(entries),
            "headers": table.headers,
        },
        "issues": dict(issues),
        "entries": entries,
    }


def analyze_category_sheet(ws) -> Dict[str, Any]:
    header_row = find_row_with_cell_value(ws, "Angebotstitel", col=2, max_rows=120)
    if not header_row:
        return {"sheet": ws.title, "type": "category", "table": None}

    table = extract_table(ws, header_row, title_col=2)
    header_to_idx = {h: i for i, h in enumerate(table.headers)}

    def get(row, key):
        idx = header_to_idx.get(key)
        if idx is None:
            return ""
        return row[idx]

    offers = []
    missing = Counter()
    brand_forms = defaultdict(set)  # lower->set(actual)

    # threshold counters
    low_photos = 0
    low_recommended_aspects = 0
    low_keywords = 0

    for row in table.rows:
        title = safe_str(get(row, "Angebotstitel"))
        brand = get(row, "Marke")
        mpn = get(row, "Herstellernummer")
        ean = get(row, "EAN")
        photos = get(row, "Anzahl der Fotos")
        rec_aspects = get(row, "Empfohlene Artikelmerkmale angegeben")
        keywords = get(row, "Anzahl der Suchbegriffe im Titel")
        item_no = get(row, "Artikelnummer")
        sku = get(row, "Bestandseinheit")

        brand_s = safe_str(brand)
        if brand_s and not is_missing(brand_s):
            brand_forms[brand_s.lower()].add(brand_s)

        # Missing tracking (treat checkmark as present)
        if is_missing(brand) and not is_present_checkmark(brand):
            missing["Marke"] += 1
        if is_missing(mpn) and not is_present_checkmark(mpn):
            missing["Herstellernummer"] += 1
        if is_missing(ean) and not is_present_checkmark(ean):
            missing["EAN"] += 1

        # numeric-ish thresholds
        p = parse_int(photos)
        if p is not None and p < 5:
            low_photos += 1
        r = parse_int(rec_aspects)
        if r is not None and r < 5:
            low_recommended_aspects += 1
        k = parse_int(keywords)
        if k is not None and k < 6:
            low_keywords += 1

        offers.append(
            {
                "sheet": ws.title,
                "title": title,
                "brand": safe_str(brand),
                "mpn": safe_str(mpn),
                "ean": safe_str(ean),
                "photos": safe_str(photos),
                "recommended_aspects_count": safe_str(rec_aspects),
                "title_keywords_count": safe_str(keywords),
                "item_number": safe_str(item_no),
                "sku": safe_str(sku),
            }
        )

    casing_issues = {
        k: sorted(list(v))
        for k, v in brand_forms.items()
        if len(v) > 1
    }

    return {
        "sheet": ws.title,
        "type": "category",
        "table": {
            "header_row": table.header_row,
            "rows": len(table.rows),
            "cols": table.last_col,
            "headers": table.headers,
        },
        "missing": dict(missing),
        "thresholds": {
            "photos_lt_5": low_photos,
            "recommended_aspects_lt_5": low_recommended_aspects,
            "title_keywords_lt_6": low_keywords,
        },
        "brand_casing_variants": casing_issues,
        "offers": offers,
    }


def main() -> None:
    parser = argparse.ArgumentParser()
    parser.add_argument("--in", dest="in_path", required=True, help="Path to input .xlsx report")
    parser.add_argument("--out-dir", dest="out_dir", required=True, help="Output directory")
    args = parser.parse_args()

    in_path = Path(args.in_path).expanduser()
    out_dir = Path(args.out_dir).expanduser()
    out_dir.mkdir(parents=True, exist_ok=True)

    wb = load_workbook(filename=str(in_path), data_only=True, read_only=True)

    overview = {
        "file": str(in_path),
        "generated_at_iso": datetime.utcnow().isoformat() + "Z",
        "sheet_names": wb.sheetnames,
        "sheets": [],
    }

    all_offers_long = []
    google_rejected_entries = []

    for name in wb.sheetnames:
        ws = wb[name]
        t = sheet_type(name)
        if t == "summary":
            overview["sheets"].append(analyze_summary_sheet(ws))
        elif t == "guide":
            overview["sheets"].append({"sheet": name, "type": "guide", "note": "Informational guide sheet (no offer table)."})
        elif t == "google_shopping_rejected":
            data = analyze_google_shopping_sheet(ws)
            overview["sheets"].append({k: v for k, v in data.items() if k != "entries"})
            google_rejected_entries.extend(data.get("entries", []))
        else:
            data = analyze_category_sheet(ws)
            overview["sheets"].append({k: v for k, v in data.items() if k != "offers"})
            all_offers_long.extend(data.get("offers", []))

    # Outputs
    write_json(out_dir / "overview.json", overview)

    # Per-category summary CSV
    summary_rows = []
    for s in overview["sheets"]:
        if s.get("type") != "category" or not s.get("table"):
            continue
        rows = s["table"]["rows"]
        missing = s.get("missing", {})
        thresholds = s.get("thresholds", {})
        summary_rows.append(
            {
                "sheet": s["sheet"],
                "offers": rows,
                "missing_brand": missing.get("Marke", 0),
                "missing_mpn": missing.get("Herstellernummer", 0),
                "missing_ean": missing.get("EAN", 0),
                "photos_lt_5": thresholds.get("photos_lt_5", 0),
                "recommended_aspects_lt_5": thresholds.get("recommended_aspects_lt_5", 0),
                "title_keywords_lt_6": thresholds.get("title_keywords_lt_6", 0),
                "brand_case_variants": len(s.get("brand_casing_variants", {}) or {}),
            }
        )

    write_csv(
        out_dir / "category_summary.csv",
        summary_rows,
        headers=[
            "sheet",
            "offers",
            "missing_brand",
            "missing_mpn",
            "missing_ean",
            "photos_lt_5",
            "recommended_aspects_lt_5",
            "title_keywords_lt_6",
            "brand_case_variants",
        ],
    )

    # Offers long CSV
    write_csv(
        out_dir / "offers_long.csv",
        all_offers_long,
        headers=[
            "sheet",
            "title",
            "brand",
            "mpn",
            "ean",
            "photos",
            "recommended_aspects_count",
            "title_keywords_count",
            "item_number",
            "sku",
        ],
    )

    # Google Shopping rejected CSV
    if google_rejected_entries:
        # Preserve original header names if possible
        all_keys = []
        seen = set()
        for row in google_rejected_entries:
            for k in row.keys():
                if k not in seen:
                    seen.add(k)
                    all_keys.append(k)
        write_csv(out_dir / "google_shopping_rejected.csv", google_rejected_entries, headers=all_keys)

    print(f"[analyze] sheets={len(wb.sheetnames)} offers_rows={len(all_offers_long)} out={out_dir}")


if __name__ == "__main__":
    main()

