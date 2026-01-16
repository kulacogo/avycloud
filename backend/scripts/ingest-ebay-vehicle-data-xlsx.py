#!/usr/bin/env python3
"""
Generate backend-readable JSON datasets from eBay-provided XLSX exports.

Inputs (repo root):
  - 2022_September_Kategorien_mit_Fahrzeugdaten_202301.xlsx
  - verpflichtende_artikelmerkmale_auto_motorradteile_24.10.23.xlsx

Outputs (backend/ebay-data):
  - vehicle-fitment-categories.json
  - required-aspects-auto-motorradteile.json

Rationale:
The Node backend intentionally avoids an XLSX parsing dependency. We ingest once and commit JSON.
"""

from __future__ import annotations

import datetime as _dt
import json
from collections import defaultdict
from pathlib import Path

from openpyxl import load_workbook


ROOT = Path(__file__).resolve().parents[2]
OUT_DIR = ROOT / "backend" / "ebay-data"

CAT_XLSX = ROOT / "2022_September_Kategorien_mit_Fahrzeugdaten_202301.xlsx"
REQ_XLSX = ROOT / "verpflichtende_artikelmerkmale_auto_motorradteile_24.10.23.xlsx"


def stamp() -> str:
    return _dt.datetime.utcnow().replace(microsecond=0).isoformat() + "Z"


def str_id(v) -> str:
    if v is None:
        return ""
    if isinstance(v, (int, float)):
        return str(int(v))
    return str(v).strip()


def main() -> None:
    OUT_DIR.mkdir(parents=True, exist_ok=True)

    # 1) Vehicle fitment categories
    wb = load_workbook(filename=str(CAT_XLSX), data_only=True, read_only=True)
    ws = wb[wb.sheetnames[0]]
    headers = [ws.cell(row=1, column=c).value for c in range(1, ws.max_column + 1)]
    headers = [str(h).strip() if h else "" for h in headers]
    idx = {h: i + 1 for i, h in enumerate(headers) if h}

    cat_col = idx["Kategorie-ID"]
    usage_col = idx["Fahrzeugverwendungsliste"]

    auto: list[str] = []
    moto: list[str] = []
    none: list[str] = []

    for r in range(2, ws.max_row + 1):
        cid = str_id(ws.cell(row=r, column=cat_col).value)
        if not cid:
            continue
        usage = (ws.cell(row=r, column=usage_col).value or "").strip().lower()
        if "autos möglich" in usage:
            auto.append(cid)
        elif "motorr" in usage and "möglich" in usage:
            moto.append(cid)
        elif "keine" in usage:
            none.append(cid)

    vehicle_json = {
        "generated_from": CAT_XLSX.name,
        "generated_at": stamp(),
        "auto_possible_category_ids": sorted(set(auto), key=lambda x: int(x)),
        "motorcycle_possible_category_ids": sorted(set(moto), key=lambda x: int(x)),
        "no_fitment_list_category_ids": sorted(set(none), key=lambda x: int(x)),
    }

    (OUT_DIR / "vehicle-fitment-categories.json").write_text(
        json.dumps(vehicle_json, ensure_ascii=False, indent=2), "utf-8"
    )

    # 2) Required aspects (auto/moto parts)
    wb2 = load_workbook(filename=str(REQ_XLSX), data_only=True, read_only=True)
    ws2 = wb2["Verpflichtende Artikelmerkmale"]
    headers = [ws2.cell(row=1, column=c).value for c in range(1, ws2.max_column + 1)]
    headers = [str(h).strip() if h else "" for h in headers]
    idx = {h: i + 1 for i, h in enumerate(headers) if h}

    id_col = idx["ID"]
    aspect_col = idx["Artikelmerkmal"]

    req = defaultdict(list)
    for r in range(2, ws2.max_row + 1):
        cid = str_id(ws2.cell(row=r, column=id_col).value)
        asp = (ws2.cell(row=r, column=aspect_col).value or "").strip()
        if not cid or not asp:
            continue
        if asp not in req[cid]:
            req[cid].append(asp)

    req_sorted = {k: req[k] for k in sorted(req.keys(), key=lambda x: int(x))}
    req_json = {
        "generated_from": REQ_XLSX.name,
        "generated_at": stamp(),
        "required_aspects_by_category_id": req_sorted,
    }
    (OUT_DIR / "required-aspects-auto-motorradteile.json").write_text(
        json.dumps(req_json, ensure_ascii=False, indent=2), "utf-8"
    )

    print("Wrote:", OUT_DIR / "vehicle-fitment-categories.json")
    print("Wrote:", OUT_DIR / "required-aspects-auto-motorradteile.json")


if __name__ == "__main__":
    main()

